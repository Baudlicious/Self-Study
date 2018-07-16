#!/usr/bin/env python
# This is x3830s python files from the book Automate Python The
# Boring Stuff with Python. 

def title(big_title, date):
    print('!!!!' + big_title.upper().center(70) + ' !!!!')
    print(date.upper().center(80))

def sep(title):
    print("-" * 80)
    print('[ + ] ' + title.title())
    print("-" * 80 +"\n")

def function(func):
    print(">> " + func + ":")

title('Working with Excel Files', '16/07/18')

import openpyxl
import os

sep('To open a Excel file use openpyxl.load_workbook()')
workbook = openpyxl.load_workbook('example.xlsx')

sep('To open a specific sheet use workbook["Sheetname"]')
sheet = workbook['Sheet1']

sep('To get a list of the names of the sheets use .sheetnames')
print(workbook.sheetnames)

sep("To read specific cells, we can use sheet['Cell'].value")
print(sheet['A1'].value)
function('Depending on what you are doing you may want to use str()')

sep("Cells can be also be specified with .cell(row=, column=)")
print(sheet.cell(row=1, column=2))
function("This is helpful if you want to loop")
for i in range(1,8):
    print(i, sheet.cell(row=i, column=2).value)

sep('We can create new spreadsheets by calling openpyxl.Workbook()')
wb = openpyxl.Workbook()
print(wb.sheetnames)
sheet = wb['Sheet']

sheet['A1'] = 42
sheet['A2'] = 'Hello'

sep('To save the sheet we use wb.save')
wb.save('example2.xlsx')

sep('To add worksheets we use create_sheet()')
sheet2 = web.create_sheet()
print(wb.sheetnames)

sep('To change title we can just assign it as a variable')
sheet2.title = 'My new sheet name'
print("sheet2.title = 'My new sheet name'")
print(wb.sheetnames)

wb.save('example3.xlsx')
sep('We can assign a sheet a specific index with create_sheet(index=0, title="title")')
wb.create_sheet(index=0, title='My other sheet')
wb.save('example4.xlsx')
